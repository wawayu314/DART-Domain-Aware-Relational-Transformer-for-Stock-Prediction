#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Run the unified baseline benchmark suite for the DART stock prediction repo.

The script keeps the original DART training pipeline untouched and writes
standalone benchmark artifacts under ``../result`` plus a Markdown summary
under ``../docs``.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import pickle
import random
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import torch
import torch.nn.functional as F
from openpyxl.utils import get_column_letter

from evaluator import evaluate
from sota_models import build_sota_model, count_parameters


ROOT = Path(__file__).resolve().parents[1]
DATASET_DIR = ROOT / "dataset"
RESULT_DIR = ROOT / "result"
DOCS_DIR = ROOT / "docs"
DEFAULT_MODELS = [
    "ALSTM",
    "Transformer",
    "PatchTST",
    "iTransformer",
    "TimesNetLite",
    "CausalStock",
    "MambaStock",
]

SENTIMENT_BASE_COLS = [
    "sentiment_polarity_mean",
    "sentiment_polarity_std",
    "sentiment_polarity_count",
    "sentiment_confidence_mean",
    "urgency_score_mean",
    "keywords_earnings_sum",
    "keywords_acquisition_sum",
    "keywords_regulation_sum",
    "keywords_partnership_sum",
    "keywords_product_sum",
    "keywords_financial_sum",
    "keywords_leadership_sum",
    "keywords_lawsuit_sum",
    "text_length_mean",
]

MODEL_FEATURE_POLICY = {
    "causalstock": {"use_sentiment": True, "use_causal_graph": True},
    "mambastock": {"use_sentiment": False, "use_causal_graph": False},
}

REFERENCE_RESULTS = {
    "SP500": [
        {
            "Model": "DART",
            "Dataset": "SP500",
            "Best Epoch": 17,
            "Test MSE": "",
            "Test IC": 0.02526211747494554,
            "Test RIC": 0.2090970209400151,
            "Test Prec@10": 0.5260869565217391,
            "Test SR": 2.616295594848619,
            "Source": "training_results.xlsx / thesis result",
        },
        {
            "Model": "Gated-DART",
            "Dataset": "SP500",
            "Best Epoch": 3,
            "Test MSE": 0.0003260114973240373,
            "Test IC": 0.03191560504778245,
            "Test RIC": 0.2422641188686827,
            "Test Prec@10": 0.5265217391304348,
            "Test SR": 2.67272027868175,
            "Source": "training_results.xlsx / thesis result",
        },
    ]
}


def get_model_policy(model_name):
    return MODEL_FEATURE_POLICY.get(
        model_name.lower(), {"use_sentiment": False, "use_causal_graph": False}
    )


def set_seed(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = False
    torch.backends.cudnn.benchmark = True


def load_market_data(market):
    market = market.upper()
    base = DATASET_DIR / market
    if not base.exists():
        raise FileNotFoundError("Dataset folder not found: {}".format(base))

    pkl_files = {
        "eod": base / "eod_data.pkl",
        "mask": base / "mask_data.pkl",
        "gt": base / "gt_data.pkl",
        "price": base / "price_data.pkl",
    }
    if all(path.exists() for path in pkl_files.values()):
        with open(pkl_files["eod"], "rb") as f:
            eod_data = pickle.load(f)
        with open(pkl_files["mask"], "rb") as f:
            mask_data = pickle.load(f)
        with open(pkl_files["gt"], "rb") as f:
            gt_data = pickle.load(f)
        with open(pkl_files["price"], "rb") as f:
            price_data = pickle.load(f)
        source = "pkl"
    else:
        npy_path = base / "{}.npy".format(market)
        if not npy_path.exists():
            raise FileNotFoundError("Cannot find pkl files or npy file for {}".format(market))
        eod_data = np.load(npy_path).astype(np.float32)
        price_data = eod_data[:, :, -1].astype(np.float32)
        mask_data = (price_data > 1e-8).astype(np.float32)
        gt_data = np.zeros_like(price_data, dtype=np.float32)
        gt_data[:, 1:] = (price_data[:, 1:] - price_data[:, :-1]) / (price_data[:, :-1] + 1e-8)
        source = "npy"

    eod_data = np.asarray(eod_data, dtype=np.float32)
    mask_data = np.asarray(mask_data, dtype=np.float32)
    gt_data = np.asarray(gt_data, dtype=np.float32)
    price_data = np.asarray(price_data, dtype=np.float32)
    if eod_data.ndim != 3:
        raise ValueError("eod_data must be [stocks, days, features], got {}".format(eod_data.shape))
    return eod_data, mask_data, gt_data, price_data, source


def load_ticker_list(market):
    market = market.upper()
    base = DATASET_DIR / market
    ticker_csv = base / "{}_ticker.csv".format(market.lower())
    if ticker_csv.exists():
        df = pd.read_csv(ticker_csv, header=None)
        return df.iloc[:, 0].astype(str).str.strip().replace("nan", np.nan).dropna().tolist()
    industry_json = base / "{}_industry_data.json".format(market.lower())
    if industry_json.exists():
        with open(industry_json, "r", encoding="utf-8") as f:
            data = json.load(f)
        return list(data.keys())
    raise FileNotFoundError("Ticker file not found for {}".format(market))


def infer_trade_dates(market, num_days):
    cache_file = DATASET_DIR / market.upper() / "cache" / "batch_1_data.pkl"
    if cache_file.exists():
        with open(cache_file, "rb") as f:
            batch_data = pickle.load(f)
        if isinstance(batch_data, list) and batch_data:
            best_df = max(batch_data, key=lambda x: len(getattr(x, "index", [])))
            index = pd.to_datetime(best_df.index)
            if len(index) >= num_days:
                return [pd.Timestamp(x).normalize() for x in index[:num_days]]
    # Fallback for markets whose cached price files have one or two missing rows.
    start = pd.Timestamp("2020-06-10")
    return list(pd.bdate_range(start=start, periods=num_days))


def build_sentiment_tensor(market, tickers, num_days):
    market = market.upper()
    cache_path = DATASET_DIR / market / "cache" / "{}_sentiment_tensor.npz".format(market.lower())
    if cache_path.exists():
        cache = np.load(cache_path, allow_pickle=True)
        tensor = cache["tensor"].astype(np.float32)
        cols = cache["columns"].tolist()
        if tensor.shape[0] == len(tickers) and tensor.shape[1] == num_days:
            return tensor, cols

    combined_path = (
        DATASET_DIR / market / "news_sentiment" / "{}_combined_daily_sentiment.csv".format(market.lower())
    )
    if not combined_path.exists():
        raise FileNotFoundError("Combined sentiment file not found: {}".format(combined_path))

    df = pd.read_csv(combined_path, usecols=["date", "ticker"] + SENTIMENT_BASE_COLS)
    for col in SENTIMENT_BASE_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    df["ticker"] = df["ticker"].astype(str).str.strip()
    df["date"] = pd.to_datetime(df["date"]).dt.tz_localize(None).dt.normalize()

    trade_dates = infer_trade_dates(market, num_days)
    date_keys = [pd.Timestamp(x).normalize() for x in trade_dates]
    date_map = {d: i for i, d in enumerate(date_keys)}
    ticker_map = {t: i for i, t in enumerate(tickers)}

    df["stock_idx"] = df["ticker"].map(ticker_map)
    df["day_idx"] = df["date"].map(date_map)
    df = df.dropna(subset=["stock_idx", "day_idx"]).copy()
    if df.empty:
        tensor = np.zeros((len(tickers), num_days, len(SENTIMENT_BASE_COLS) + 4), dtype=np.float32)
        cols = SENTIMENT_BASE_COLS + [
            "market_sentiment_mean",
            "relative_sentiment",
            "market_news_volume",
            "relative_news_volume",
        ]
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        np.savez_compressed(cache_path, tensor=tensor, columns=np.asarray(cols, dtype=object))
        return tensor, cols

    df["stock_idx"] = df["stock_idx"].astype(np.int64)
    df["day_idx"] = df["day_idx"].astype(np.int64)
    tensor = np.zeros((len(tickers), num_days, len(SENTIMENT_BASE_COLS)), dtype=np.float32)
    rows = df[["stock_idx", "day_idx"] + SENTIMENT_BASE_COLS].to_numpy()
    stock_idx = rows[:, 0].astype(np.int64)
    day_idx = rows[:, 1].astype(np.int64)
    for feat_idx in range(len(SENTIMENT_BASE_COLS)):
        tensor[stock_idx, day_idx, feat_idx] = rows[:, 2 + feat_idx].astype(np.float32)

    pol_idx = SENTIMENT_BASE_COLS.index("sentiment_polarity_mean")
    cnt_idx = SENTIMENT_BASE_COLS.index("sentiment_polarity_count")
    polarity = tensor[:, :, pol_idx]
    news_count = tensor[:, :, cnt_idx]
    coverage = (news_count > 0).astype(np.float32)
    denom = np.clip(coverage.sum(axis=0, keepdims=True), 1.0, None)
    market_sentiment = (polarity * coverage).sum(axis=0, keepdims=True) / denom
    market_volume = news_count.sum(axis=0, keepdims=True) / denom
    relative_sentiment = polarity - market_sentiment
    relative_volume = np.log1p(news_count) - np.log1p(market_volume)
    extra = np.stack(
        [
            np.repeat(market_sentiment, len(tickers), axis=0),
            relative_sentiment,
            np.repeat(market_volume, len(tickers), axis=0),
            relative_volume,
        ],
        axis=-1,
    ).astype(np.float32)
    full_tensor = np.concatenate([tensor, extra], axis=-1).astype(np.float32)
    cols = SENTIMENT_BASE_COLS + [
        "market_sentiment_mean",
        "relative_sentiment",
        "market_news_volume",
        "relative_news_volume",
    ]
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(cache_path, tensor=full_tensor, columns=np.asarray(cols, dtype=object))
    return full_tensor, cols


def masked_zscore(array, mask):
    masked = array * mask
    denom = np.clip(mask.sum(axis=1, keepdims=True), 1.0, None)
    mean = masked.sum(axis=1, keepdims=True) / denom
    var = (((array - mean) * mask) ** 2).sum(axis=1, keepdims=True) / denom
    std = np.sqrt(np.clip(var, 1e-6, None))
    return ((array - mean) / std) * mask


def build_causal_graph(market, gt_data, mask_data, sentiment_tensor, train_end_index, topk=8, max_lag=3):
    market = market.upper()
    cache_path = (
        DATASET_DIR
        / market
        / "cache"
        / "causalstock_graph_t{}_k{}_l{}.npz".format(train_end_index, topk, max_lag)
    )
    if cache_path.exists():
        cache = np.load(cache_path)
        edge_index = torch.from_numpy(cache["edge_index"]).long()
        edge_weight = torch.from_numpy(cache["edge_weight"]).float()
        return edge_index, edge_weight

    train_returns = gt_data[:, :train_end_index].astype(np.float32)
    train_mask = mask_data[:, :train_end_index].astype(np.float32)
    sentiment_signal = sentiment_tensor[:, :train_end_index, 0].astype(np.float32)

    ret_z = masked_zscore(train_returns, train_mask)
    sent_z = masked_zscore(sentiment_signal, train_mask)
    stock_num = train_returns.shape[0]
    scores = np.zeros((stock_num, stock_num), dtype=np.float32)

    for lag in range(1, max_lag + 1):
        src_ret = ret_z[:, :-lag]
        tgt_ret = ret_z[:, lag:]
        src_sent = sent_z[:, :-lag]
        length = max(src_ret.shape[1], 1)
        ret_score = tgt_ret @ src_ret.T / float(length)
        sent_score = tgt_ret @ src_sent.T / float(length)
        scores += (0.7 * ret_score + 0.3 * sent_score) / float(lag)

    np.fill_diagonal(scores, 0.0)
    edge_src = []
    edge_dst = []
    edge_weight = []
    for dst in range(stock_num):
        row = scores[dst]
        if topk >= stock_num:
            candidates = np.argsort(-np.abs(row))
        else:
            candidates = np.argpartition(-np.abs(row), topk)[:topk]
            candidates = candidates[np.argsort(-np.abs(row[candidates]))]
        candidates = [int(src) for src in candidates if src != dst and abs(float(row[src])) > 1e-6][:topk]
        if not candidates:
            continue
        weights = row[candidates].astype(np.float32)
        norm = np.sum(np.abs(weights)) + 1e-6
        weights = weights / norm
        for src, weight in zip(candidates, weights):
            edge_src.append(src)
            edge_dst.append(dst)
            edge_weight.append(weight)

    edge_index = np.asarray([edge_src, edge_dst], dtype=np.int64)
    edge_weight = np.asarray(edge_weight, dtype=np.float32)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(cache_path, edge_index=edge_index, edge_weight=edge_weight)
    return torch.from_numpy(edge_index).long(), torch.from_numpy(edge_weight).float()


def compute_split_indices(trade_days, lookback, steps, train_ratio=0.6, valid_ratio=0.2):
    min_valid_days = max(50, lookback + steps + 10)
    min_test_days = max(50, lookback + steps + 10)
    min_train_days = max(100, lookback + steps + 20)
    valid_index = max(min_train_days, math.floor(trade_days * train_ratio))
    test_index = max(valid_index + min_valid_days, math.floor(trade_days * (train_ratio + valid_ratio)))
    if trade_days - test_index < min_test_days:
        test_index = trade_days - min_test_days
        if test_index <= valid_index:
            available_days = trade_days - min_train_days
            if available_days >= min_valid_days + min_test_days:
                valid_index = min_train_days
                test_index = trade_days - min_test_days
            else:
                raise ValueError("Dataset too small: {} days".format(trade_days))
    if valid_index >= test_index or test_index >= trade_days:
        raise ValueError(
            "Invalid split: valid_index={}, test_index={}, trade_days={}".format(
                valid_index, test_index, trade_days
            )
        )
    return valid_index, test_index


def standardize_by_train(eod_data, valid_index):
    train = eod_data[:, :valid_index, :]
    mean = train.mean(axis=(0, 1), keepdims=True)
    std = train.std(axis=(0, 1), keepdims=True)
    std[std == 0] = 1.0
    data = (eod_data - mean) / std
    data = np.nan_to_num(data, nan=0.0, posinf=0.0, neginf=0.0).astype(np.float32)
    return data


def make_offsets(start_index, end_index, lookback, steps):
    start_offset = start_index - lookback - steps + 1
    end_offset = end_index - lookback - steps + 1
    offsets = np.arange(start_offset, end_offset, dtype=np.int64)
    offsets = offsets[offsets >= 0]
    return offsets


def get_batch(eod_data, mask_data, gt_data, offset, lookback, steps, device):
    target_day = offset + lookback + steps - 1
    x = eod_data[:, offset : offset + lookback, :]
    mask_window = mask_data[:, offset : offset + lookback + steps]
    mask = np.min(mask_window, axis=1, keepdims=True)
    y = gt_data[:, target_day : target_day + 1]
    x_t = torch.from_numpy(x).float().to(device)
    y_t = torch.from_numpy(y).float().to(device)
    mask_t = torch.from_numpy(mask).float().to(device)
    return x_t, y_t, mask_t, target_day


def masked_mse(prediction, target, mask):
    denom = torch.clamp(mask.sum(), min=1.0)
    return torch.sum(((prediction - target) * mask) ** 2) / denom


def pairwise_rank_loss(prediction, target, mask):
    pred_diff = prediction - prediction.t()
    target_diff = target - target.t()
    pair_mask = mask @ mask.t()
    denom = torch.clamp(pair_mask.sum(), min=1.0)
    loss = F.relu(-pred_diff * target_diff) * pair_mask
    return torch.sum(loss) / denom


def combined_loss(prediction, target, mask, alpha):
    reg_loss = masked_mse(prediction, target, mask)
    rank_loss = pairwise_rank_loss(prediction, target, mask)
    return reg_loss + alpha * rank_loss, reg_loss, rank_loss


def evaluate_model(model, eod_data, mask_data, gt_data, offsets, start_index, end_index, lookback, steps, device):
    model.eval()
    stock_num = eod_data.shape[0]
    num_days = end_index - start_index
    pred_mat = np.zeros((stock_num, num_days), dtype=np.float32)
    gt_mat = np.zeros((stock_num, num_days), dtype=np.float32)
    mask_mat = np.zeros((stock_num, num_days), dtype=np.float32)
    total_loss = 0.0
    total_reg = 0.0
    total_rank = 0.0

    with torch.no_grad():
        for offset in offsets:
            x, y, mask, target_day = get_batch(eod_data, mask_data, gt_data, int(offset), lookback, steps, device)
            pred = model(x)
            loss, reg_loss, rank_loss = combined_loss(pred, y, mask, alpha=0.0)
            col = target_day - start_index
            if 0 <= col < num_days:
                pred_mat[:, col] = pred[:, 0].detach().cpu().numpy()
                gt_mat[:, col] = y[:, 0].detach().cpu().numpy()
                mask_mat[:, col] = mask[:, 0].detach().cpu().numpy()
            total_loss += float(loss.item())
            total_reg += float(reg_loss.item())
            total_rank += float(rank_loss.item())

    perf = evaluate(pred_mat, gt_mat, mask_mat)
    count = max(len(offsets), 1)
    perf["loss"] = total_loss / count
    perf["reg_loss"] = total_reg / count
    perf["rank_loss"] = total_rank / count
    return perf


def clean_metric(value):
    try:
        value = float(value)
        if np.isnan(value) or np.isinf(value):
            return 0.0
        return value
    except Exception:
        return 0.0


def prepare_model_inputs(model_name, args, eod_data, mask_data, gt_data, valid_index):
    model_policy = get_model_policy(model_name)
    feature_tensor = np.asarray(eod_data, dtype=np.float32)
    model_kwargs = {}
    notes = {"feature_dim": int(feature_tensor.shape[2]), "uses_sentiment": False, "causal_edges": 0}

    if model_policy["use_sentiment"]:
        tickers = load_ticker_list(args.market)
        sentiment_tensor, sentiment_cols = build_sentiment_tensor(
            args.market, tickers=tickers[: feature_tensor.shape[0]], num_days=feature_tensor.shape[1]
        )
        sentiment_tensor = standardize_by_train(sentiment_tensor, valid_index)
        feature_tensor = np.concatenate([feature_tensor, sentiment_tensor], axis=-1).astype(np.float32)
        model_kwargs["price_input_dim"] = int(eod_data.shape[2])
        model_kwargs["news_input_dim"] = int(sentiment_tensor.shape[2])
        notes["uses_sentiment"] = True
        notes["sentiment_cols"] = sentiment_cols

        if model_policy["use_causal_graph"]:
            edge_index, edge_weight = build_causal_graph(
                args.market,
                gt_data=gt_data,
                mask_data=mask_data,
                sentiment_tensor=sentiment_tensor,
                train_end_index=valid_index,
                topk=8,
                max_lag=3,
            )
            model_kwargs["edge_index"] = edge_index
            model_kwargs["edge_weight"] = edge_weight
            notes["causal_edges"] = int(edge_weight.numel())

    notes["feature_dim"] = int(feature_tensor.shape[2])
    return feature_tensor, model_kwargs, notes


def train_one_model(model_name, args, data_pack, split_pack, device):
    feature_data, mask_data, gt_data, model_kwargs, input_notes = data_pack
    valid_index, test_index, train_offsets, valid_offsets, test_offsets = split_pack

    set_seed(args.seed)
    model = build_sota_model(
        model_name,
        input_dim=feature_data.shape[2],
        lookback=args.lookback,
        **model_kwargs
    ).to(device)
    optimizer = torch.optim.AdamW(model.parameters(), lr=args.lr, weight_decay=args.weight_decay)
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=max(args.epochs, 1), eta_min=args.min_lr)

    best_valid_metric = -float("inf")
    best_valid_perf = None
    best_test_perf = None
    best_epoch = 0
    no_improve = 0

    print("\n=== Training {} ===".format(model_name))
    print("Parameters: {:,}".format(count_parameters(model)))

    for epoch in range(1, args.epochs + 1):
        model.train()
        shuffled = np.array(train_offsets, copy=True)
        np.random.shuffle(shuffled)
        total_loss = 0.0
        total_reg = 0.0
        total_rank = 0.0
        seen = 0

        for offset in shuffled:
            x, y, mask, _ = get_batch(feature_data, mask_data, gt_data, int(offset), args.lookback, args.steps, device)
            optimizer.zero_grad(set_to_none=True)
            pred = model(x)
            loss, reg_loss, rank_loss = combined_loss(pred, y, mask, args.alpha)
            loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), args.grad_clip)
            optimizer.step()

            total_loss += float(loss.item())
            total_reg += float(reg_loss.item())
            total_rank += float(rank_loss.item())
            seen += 1

        scheduler.step()
        valid_perf = evaluate_model(
            model, feature_data, mask_data, gt_data, valid_offsets, valid_index, test_index,
            args.lookback, args.steps, device
        )
        test_perf = evaluate_model(
            model, feature_data, mask_data, gt_data, test_offsets, test_index, mask_data.shape[1],
            args.lookback, args.steps, device
        )
        metric = clean_metric(valid_perf.get(args.select_metric, 0.0))

        print(
            "Epoch {}/{} loss={:.4e} valid_IC={:.4f} valid_SR={:.4f} "
            "test_IC={:.4f} test_SR={:.4f}".format(
                epoch,
                args.epochs,
                total_loss / max(seen, 1),
                clean_metric(valid_perf.get("IC")),
                clean_metric(valid_perf.get("sharpe5")),
                clean_metric(test_perf.get("IC")),
                clean_metric(test_perf.get("sharpe5")),
            )
        )

        if metric > best_valid_metric:
            best_valid_metric = metric
            best_valid_perf = dict(valid_perf)
            best_test_perf = dict(test_perf)
            best_epoch = epoch
            no_improve = 0
        else:
            no_improve += 1
            if no_improve >= args.patience:
                print("Early stopping {} at epoch {}".format(model_name, epoch))
                break

    row = {
        "Model": model_name,
        "Dataset": args.market.upper(),
        "Best Epoch": best_epoch,
        "Valid MSE": clean_metric(best_valid_perf.get("mse")) if best_valid_perf else 0.0,
        "Valid IC": clean_metric(best_valid_perf.get("IC")) if best_valid_perf else 0.0,
        "Valid RIC": clean_metric(best_valid_perf.get("RIC")) if best_valid_perf else 0.0,
        "Valid Prec@10": clean_metric(best_valid_perf.get("prec_10")) if best_valid_perf else 0.0,
        "Valid SR": clean_metric(best_valid_perf.get("sharpe5")) if best_valid_perf else 0.0,
        "Test MSE": clean_metric(best_test_perf.get("mse")) if best_test_perf else 0.0,
        "Test IC": clean_metric(best_test_perf.get("IC")) if best_test_perf else 0.0,
        "Test RIC": clean_metric(best_test_perf.get("RIC")) if best_test_perf else 0.0,
        "Test Prec@10": clean_metric(best_test_perf.get("prec_10")) if best_test_perf else 0.0,
        "Test SR": clean_metric(best_test_perf.get("sharpe5")) if best_test_perf else 0.0,
        "Lookback": args.lookback,
        "Epochs": args.epochs,
        "Learning Rate": args.lr,
        "Alpha": args.alpha,
        "Params": count_parameters(model),
        "Input Dim": int(feature_data.shape[2]),
        "Uses Sentiment": bool(input_notes.get("uses_sentiment", False)),
        "Causal Edges": int(input_notes.get("causal_edges", 0)),
        "Selection Metric": args.select_metric,
        "Run Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return row


def write_csv(rows, path):
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(rows, path):
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOTA Benchmark"
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col_idx, column_cells in enumerate(ws.columns):
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = min(width + 2, 28)
    wb.save(path)


def fmt(value, digits=4):
    if value == "":
        return "-"
    try:
        return "{:.{}f}".format(float(value), digits)
    except Exception:
        return str(value)


def markdown_table(rows):
    lines = [
        "| Model | Dataset | Best Epoch | IC | RIC | Prec@10 | SR |",
        "|---|---|---:|---:|---:|---:|---:|",
    ]
    for row in rows:
        lines.append(
            "| {} | {} | {} | {} | {} | {} | {} |".format(
                row.get("Model", ""),
                row.get("Dataset", ""),
                row.get("Best Epoch", ""),
                fmt(row.get("Test IC", "")),
                fmt(row.get("Test RIC", "")),
                fmt(row.get("Test Prec@10", "")),
                fmt(row.get("Test SR", "")),
            )
        )
    return "\n".join(lines)


def write_markdown_summary(rows, args, path):
    market = args.market.upper()
    reference_rows = REFERENCE_RESULTS.get(market, [])
    all_rows = list(rows) + reference_rows
    path.parent.mkdir(parents=True, exist_ok=True)
    rerun_cmd = (
        "python src/sota_benchmark.py --market {} --epochs {} --patience {} "
        "--lookback {} --models {}"
    ).format(market, args.epochs, args.patience, args.lookback, ",".join(args.models))

    lines = [
        "# Unified Baseline Benchmark",
        "",
        "Generated by `src/sota_benchmark.py`.",
        "",
        "## Settings",
        "",
        "- Market: `{}`".format(market),
        "- Models: `{}`".format(", ".join(args.models)),
        "- Lookback: `{}`".format(args.lookback),
        "- Steps ahead: `{}`".format(args.steps),
        "- Epochs: `{}`".format(args.epochs),
        "- Selection metric: `{}`".format(args.select_metric),
        "- Seed: `{}`".format(args.seed),
        "",
        "## Test Metrics",
        "",
        markdown_table(all_rows),
        "",
        "## Notes",
        "",
        "- Supported comparison models: `ALSTM`, `Transformer`, `PatchTST`, `iTransformer`, `TimesNetLite`, `CausalStock`, and `MambaStock`.",
        "- All models in this benchmark are evaluated under the same split protocol and metric set for the selected market.",
        "- Models that need extra aligned inputs read them from the same dataset layout used by the main experiments.",
        "- `DART` and `Gated-DART` rows are included as reference results when they are available for the selected market.",
        "",
        "## Reproduce",
        "",
        "```bash",
        rerun_cmd,
        "```",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Run the unified baseline benchmark suite used for DART comparisons."
    )
    parser.add_argument("--market", default="SP500", help="Dataset name: SP500, NASDAQ, NYSE, A_SHARE")
    parser.add_argument("--models", default=",".join(DEFAULT_MODELS), help="Comma-separated model names or 'all'")
    parser.add_argument("--lookback", type=int, default=32)
    parser.add_argument("--steps", type=int, default=1)
    parser.add_argument("--epochs", type=int, default=12)
    parser.add_argument("--patience", type=int, default=5)
    parser.add_argument("--lr", type=float, default=2e-4)
    parser.add_argument("--min-lr", type=float, default=1e-6)
    parser.add_argument("--weight-decay", type=float, default=1e-3)
    parser.add_argument("--alpha", type=float, default=0.5)
    parser.add_argument("--grad-clip", type=float, default=1.0)
    parser.add_argument("--seed", type=int, default=2026)
    parser.add_argument("--device", default="auto", choices=["auto", "cpu", "cuda"])
    parser.add_argument("--select-metric", default="IC", choices=["IC", "RIC", "prec_10", "sharpe5"])
    parser.add_argument("--output-prefix", default=None)
    args = parser.parse_args()
    if args.models.lower() == "all":
        args.models = list(DEFAULT_MODELS)
    else:
        args.models = [m.strip() for m in args.models.split(",") if m.strip()]
    return args


def main():
    args = parse_args()
    set_seed(args.seed)

    if args.device == "cuda" or (args.device == "auto" and torch.cuda.is_available()):
        device = torch.device("cuda")
    else:
        device = torch.device("cpu")

    print("Device: {}".format(device))
    print("Market: {}".format(args.market.upper()))
    print("Models: {}".format(", ".join(args.models)))

    eod_data, mask_data, gt_data, price_data, source = load_market_data(args.market)
    valid_index, test_index = compute_split_indices(mask_data.shape[1], args.lookback, args.steps)
    eod_data = standardize_by_train(eod_data, valid_index)

    train_offsets = make_offsets(args.lookback + args.steps - 1, valid_index, args.lookback, args.steps)
    valid_offsets = make_offsets(valid_index, test_index, args.lookback, args.steps)
    test_offsets = make_offsets(test_index, mask_data.shape[1], args.lookback, args.steps)
    split_pack = (valid_index, test_index, train_offsets, valid_offsets, test_offsets)
    print("Data source: {}, shape={}, valid_index={}, test_index={}".format(source, eod_data.shape, valid_index, test_index))
    print("Offsets: train={}, valid={}, test={}".format(len(train_offsets), len(valid_offsets), len(test_offsets)))

    rows = []
    model_meta = {}
    for model_name in args.models:
        feature_data, model_kwargs, input_notes = prepare_model_inputs(
            model_name, args, eod_data, mask_data, gt_data, valid_index
        )
        data_pack = (feature_data, mask_data, gt_data, model_kwargs, input_notes)
        rows.append(train_one_model(model_name, args, data_pack, split_pack, device))
        model_meta[model_name] = input_notes

    prefix = args.output_prefix or "sota_benchmark_{}".format(args.market.lower())
    csv_path = RESULT_DIR / "{}.csv".format(prefix)
    xlsx_path = RESULT_DIR / "{}.xlsx".format(prefix)
    json_path = RESULT_DIR / "{}_meta.json".format(prefix)
    summary_path = DOCS_DIR / "{}.md".format(prefix)

    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)
    meta = {
        "market": args.market.upper(),
        "models": args.models,
        "lookback": args.lookback,
        "steps": args.steps,
        "epochs": args.epochs,
        "patience": args.patience,
        "seed": args.seed,
        "select_metric": args.select_metric,
        "data_shape": list(eod_data.shape),
        "valid_index": int(valid_index),
        "test_index": int(test_index),
        "model_inputs": model_meta,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    json_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown_summary(rows, args, summary_path)

    print("\nSaved:")
    print("  {}".format(csv_path))
    print("  {}".format(xlsx_path))
    print("  {}".format(json_path))
    print("  {}".format(summary_path))
    print("\nSummary:")
    print(markdown_table(rows + REFERENCE_RESULTS.get(args.market.upper(), [])))


if __name__ == "__main__":
    main()
