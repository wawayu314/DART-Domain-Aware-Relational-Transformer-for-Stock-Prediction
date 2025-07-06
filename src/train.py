import random
import numpy as np
import os
import torch as torch
import torch.nn as nn # Added for activation function classes
from load_data import load_EOD_data
from evaluator import evaluate
from model import get_loss, StockPredict
import pickle
import openpyxl # æ–°å¢å¯¼å…¥
from openpyxl.utils import get_column_letter # æ–°å¢å¯¼å…¥
import math # æ–°å¢å¯¼å…¥ï¼Œç”¨äºè®¡ç®—åˆ†å‰²ç‚¹
import pandas as pd # æ–°å¢å¯¼å…¥ï¼Œç”¨äºåŠ è½½è‚¡ç¥¨ä»£ç åˆ—è¡¨
import json # æ–°å¢å¯¼å…¥ï¼Œç”¨äºåŠ è½½è¡Œä¸šæ•°æ®


def load_ticker_list(market_name: str):
   
    try:
        # å°è¯•åŠ è½½tickeræ–‡ä»¶
        ticker_path = f'../dataset/{market_name}/{market_name.lower()}_ticker.csv'
        if os.path.exists(ticker_path):
            df = pd.read_csv(ticker_path)
            # å‡è®¾ç¬¬ä¸€åˆ—æ˜¯è‚¡ç¥¨ä»£ç 
            ticker_column = df.columns[0]
            ticker_list = df[ticker_column].tolist()
            print(f"âœ… æˆåŠŸä» {ticker_path} åŠ è½½è‚¡ç¥¨åˆ—è¡¨: {len(ticker_list)} åªè‚¡ç¥¨")
            return ticker_list
        
        # å¦‚æœæ²¡æœ‰tickeræ–‡ä»¶ï¼Œå°è¯•ä»è¡Œä¸šæ•°æ®ä¸­æå–
        industry_path = f'../dataset/{market_name}/{market_name.lower()}_industry_data.json'
        if os.path.exists(industry_path):
            with open(industry_path, 'r', encoding='utf-8') as f:
                industry_data = json.load(f)
            ticker_list = list(industry_data.keys())
            print(f"âœ… ä»è¡Œä¸šæ•°æ®æ–‡ä»¶æå–è‚¡ç¥¨åˆ—è¡¨: {len(ticker_list)} åªè‚¡ç¥¨")
            return ticker_list
            
    except Exception as e:
        print(f"âŒ åŠ è½½è‚¡ç¥¨åˆ—è¡¨å¤±è´¥: {e}")
    
    # å¦‚æœéƒ½å¤±è´¥äº†ï¼Œè¿”å›ç©ºåˆ—è¡¨
    print(f"âš ï¸ è­¦å‘Š: æ— æ³•åŠ è½½ {market_name} çš„è‚¡ç¥¨åˆ—è¡¨ï¼Œç¨€ç–æ³¨æ„åŠ›å°†æ— æ³•ä½¿ç”¨")
    return []


np.random.seed(123456789)
torch.random.manual_seed(12345678)
device = torch.device("cuda") if torch.cuda.is_available() else 'cpu' # å¦‚æœCUDAå¯ç”¨åˆ™ä½¿ç”¨GPUï¼Œå¦åˆ™ä½¿ç”¨CPU

data_path = '../dataset' # æ•°æ®è·¯å¾„
market_name = 'SP500' # æ˜ç¡®è®¾ç½®ä¸ºSP500å¸‚åœº
relation_name = 'wikidata' # å…³ç³»åç§° (ä¼¼ä¹æœªä½¿ç”¨)
stock_num = 1026 # è‚¡ç¥¨æ•°é‡ (æ­¤å€¼å°†åœ¨æ•°æ®åŠ è½½åè¢«åŠ¨æ€è¦†ç›–)

# ============================================
# ğŸ¯ æ ¹æ®ä¸åŒå¸‚åœºé…ç½®å‚æ•°
# ============================================
if market_name.upper() == 'NYSE':
    # ğŸ›ï¸ çº½çº¦è¯åˆ¸äº¤æ˜“æ‰€ - æˆç†Ÿå¸‚åœºé…ç½®
    lookback_length = 32
    epochs = 60
    learning_rate = 0.00005
    alpha = 0.75  # è¾ƒä½çš„alphaï¼Œé‡è§†æ’åæŸå¤±
    scale_factor = 3
    activation_str = 'GELU'
    attention_heads = 8
    attention_dropout = 0.1
    weight_decay = 0.0001
    attention_ffn_dim_multiplier = 6
    print("ğŸ›ï¸ ä½¿ç”¨NYSEå¸‚åœºé…ç½® - æˆç†Ÿå¸‚åœºï¼Œæ³¨é‡ç¨³å®šæ€§å’Œæ’åè´¨é‡")
    
elif market_name.upper() == 'NASDAQ':
    # ğŸš€ çº³æ–¯è¾¾å…‹ - ç§‘æŠ€è‚¡é…ç½®
    lookback_length = 64
    epochs = 60
    learning_rate = 0.0001
    alpha = 0.8  # æ›´ä½çš„alphaï¼Œç§‘æŠ€è‚¡æ³¢åŠ¨å¤§éœ€è¦æ›´å¥½çš„æ’å
    scale_factor = 3
    activation_str = 'Hardswish'
    attention_heads = 5
    attention_dropout = 0.5
    weight_decay = 0.0005
    attention_ffn_dim_multiplier = 4
    print("ğŸš€ ä½¿ç”¨NASDAQå¸‚åœºé…ç½® - ç§‘æŠ€è‚¡å¸‚åœºï¼Œæ›´é«˜å¤æ‚åº¦æ¨¡å‹é€‚åº”å¿«é€Ÿå˜åŒ–")
    
elif market_name.upper() == 'SP500':
    # ğŸ“ˆ æ ‡æ™®500 - å¤§ç›˜è“ç­¹é…ç½®
    lookback_length = 32
    epochs = 60
    learning_rate = 0.00002
    alpha = 0.5  # é€‚ä¸­çš„alphaï¼Œå¹³è¡¡å›å½’å’Œæ’å
    scale_factor = 4
    activation_str = 'GELU'
    attention_heads = 6
    attention_dropout = 0.2
    weight_decay = 4e-4
    attention_ffn_dim_multiplier = 4
    print("ğŸ“ˆ ä½¿ç”¨SP500å¸‚åœºé…ç½® - å¤§ç›˜è“ç­¹ï¼Œé•¿æœŸè¶‹åŠ¿æ˜æ˜¾ï¼Œéœ€è¦æ›´é•¿å›çœ‹çª—å£")
    
elif market_name.upper() == 'A_SHARE':
    # ğŸ‡¨ğŸ‡³ Aè‚¡å¸‚åœº - ç‰¹æ®Šé…ç½®
    lookback_length = 36
    epochs = 150
    learning_rate = 8e-5
    alpha = 0.2   # æœ€ä½çš„alphaï¼ŒAè‚¡å¸‚åœºéœ€è¦ç‰¹åˆ«æ³¨é‡æ’å
    scale_factor = 1
    activation_str = 'GELU'
    attention_heads = 4
    attention_dropout = 0.35
    weight_decay = 8e-4
    attention_ffn_dim_multiplier = 3
    print("ğŸ‡¨ğŸ‡³ ä½¿ç”¨A_SHAREå¸‚åœºé…ç½® - Aè‚¡å¸‚åœºï¼Œé«˜æ³¢åŠ¨æ€§ï¼Œéœ€è¦æ›´å¼ºçš„æ­£åˆ™åŒ–")
    
else:
    # ğŸ”§ é»˜è®¤é…ç½® - é€šç”¨è®¾ç½®
    lookback_length = 48
    epochs = 80
    learning_rate = 5e-5
    alpha = 0.3
    scale_factor = 2
    activation_str = 'GELU'
    attention_heads = 6
    attention_dropout = 0.25
    weight_decay = 5e-4
    attention_ffn_dim_multiplier = 4
    print("ğŸ”§ ä½¿ç”¨é»˜è®¤é…ç½® - é€šç”¨è®¾ç½®ï¼Œé€‚ç”¨äºæœªç‰¹åˆ«ä¼˜åŒ–çš„å¸‚åœº")

# å›ºå®šå‚æ•°
fea_num = 5 # ç‰¹å¾æ•°é‡
steps = 1 # é¢„æµ‹æ­¥é•¿

print(f"ğŸ“Š {market_name} å¸‚åœºå‚æ•°é…ç½®:")
print(f"   å›çœ‹é•¿åº¦: {lookback_length} å¤©")
print(f"   è®­ç»ƒè½®æ•°: {epochs} epochs")
print(f"   å­¦ä¹ ç‡: {learning_rate:.1e}")
print(f"   Alpha (æŸå¤±æƒé‡): {alpha}")
print(f"   æ¨¡å‹å¤æ‚åº¦: scale_factor={scale_factor}")
print(f"   æ³¨æ„åŠ›å¤´æ•°: {attention_heads}")
print(f"   æ³¨æ„åŠ›dropout: {attention_dropout}")
print(f"   æƒé‡è¡°å‡: {weight_decay:.1e}")
print(f"   FFNç»´åº¦ä¹˜å­: {attention_ffn_dim_multiplier}")
print("="*60)

# ç¡®å®šæ¿€æ´»å‡½æ•°ç±»
if activation_str == 'GELU':
    activation_fn_to_pass = nn.GELU
elif activation_str == 'Hardswish': 
    activation_fn_to_pass = nn.Hardswish
elif activation_str == 'ReLU':
    activation_fn_to_pass = nn.ReLU
else:
    # å¦‚æœæŒ‡å®šäº†ä¸æ”¯æŒçš„æ¿€æ´»å‡½æ•°ï¼Œé»˜è®¤ä½¿ç”¨GELUæˆ–æŠ¥é”™
    print(f"Warning: Unsupported activation '{activation_str}', defaulting to GELU.")
    activation_fn_to_pass = nn.GELU

dataset_path = '../dataset/' + market_name # æ•°æ®é›†å®Œæ•´è·¯å¾„
if market_name == "SP500":
    # SP500æ•°æ®é›†çš„ç‰¹æ®ŠåŠ è½½å’Œé¢„å¤„ç†é€»è¾‘
    data = np.load('../dataset/SP500/SP500.npy')
    print(f"åŸå§‹SP500æ•°æ®å½¢çŠ¶: {data.shape} (è‚¡ç¥¨æ•°: {data.shape[0]}, å¤©æ•°: {data.shape[1]}, ç‰¹å¾æ•°: {data.shape[2]})")
    
    # ä½¿ç”¨å…¨éƒ¨5å¹´æ•°æ®ï¼Œä¸å†è¿›è¡Œåˆ‡ç‰‡
    print(f"ä½¿ç”¨å…¨éƒ¨5å¹´æ•°æ® (2020-2024)ï¼Œæ•°æ®å½¢çŠ¶: {data.shape}")
    
    price_data = data[:, :, -1] # ä»·æ ¼æ•°æ®
    mask_data = np.ones((data.shape[0], data.shape[1])) # æ©ç æ•°æ®
    eod_data = data # EODæ•°æ®
    gt_data = np.zeros((data.shape[0], data.shape[1])) # çœŸå®æ”¶ç›Šç‡æ•°æ®åˆå§‹åŒ–
    # è®¡ç®—çœŸå®æ”¶ç›Šç‡
    for ticket in range(0, data.shape[0]):
        for row in range(1, data.shape[1]):
            gt_data[ticket][row] = (data[ticket][row][-1] - data[ticket][row - steps][-1]) / \
                                   (data[ticket][row - steps][-1] + 1e-8) # æ·»åŠ å°å€¼é¿å…é™¤é›¶
else:
    # å…¶ä»–å¸‚åœº (å¦‚NASDAQ, NYSE) çš„æ•°æ®åŠ è½½é€»è¾‘
    try:
        with open(os.path.join(dataset_path, "eod_data.pkl"), "rb") as f:
            eod_data = pickle.load(f)
        with open(os.path.join(dataset_path, "mask_data.pkl"), "rb") as f:
            mask_data = pickle.load(f)
        with open(os.path.join(dataset_path, "gt_data.pkl"), "rb") as f:
            gt_data = pickle.load(f)
        with open(os.path.join(dataset_path, "price_data.pkl"), "rb") as f:
            price_data = pickle.load(f)
    except FileNotFoundError as e:
        print(f"Error loading data files: {e}")
        print(f"Please ensure {dataset_path}/ contains eod_data.pkl, mask_data.pkl, gt_data.pkl, price_data.pkl")
        exit()

# Verify stock_num if data is loaded (e.g., from eod_data shape)
if eod_data.shape[0] != stock_num:
    print(f"Warning: stock_num ({stock_num}) does not match eod_data.shape[0] ({eod_data.shape[0]}). Adjusting stock_num.")
    stock_num = eod_data.shape[0]

trade_dates = mask_data.shape[1]

# --- é€šç”¨è®¾ç½®å’ŒåŠ¨æ€åˆ†å‰² ---
print(f"å¸‚åœºä¸º {market_name}ï¼Œå®é™…åŠ è½½è‚¡ç¥¨æ•°é‡æ›´æ–°ä¸º: {stock_num}")
print(f"æ€»äº¤æ˜“å¤©æ•°: {trade_dates}")

# åŠ¨æ€è®¡ç®—æ•°æ®é›†åˆ†å‰²ç‚¹
train_ratio = 0.60
valid_ratio = 0.20
# æµ‹è¯•é›†æ¯”ä¾‹ç”± (1 - train_ratio - valid_ratio) éšå¼ç¡®å®š

# ç¡®ä¿æœ€å°æ•°æ®é›†å¤§å°
min_valid_days = max(50, lookback_length + steps + 10)  # éªŒè¯é›†æœ€å°‘50å¤©æˆ–lookback+steps+10å¤©
min_test_days = max(50, lookback_length + steps + 10)   # æµ‹è¯•é›†æœ€å°‘50å¤©æˆ–lookback+steps+10å¤©
min_train_days = max(100, lookback_length + steps + 20) # è®­ç»ƒé›†æœ€å°‘100å¤©æˆ–lookback+steps+20å¤©

# è®¡ç®—åˆ†å‰²ç‚¹
valid_index = max(min_train_days, math.floor(trade_dates * train_ratio))
test_index = max(valid_index + min_valid_days, math.floor(trade_dates * (train_ratio + valid_ratio)))

# ç¡®ä¿æµ‹è¯•é›†ä¹Ÿæœ‰è¶³å¤Ÿçš„æ•°æ®
if trade_dates - test_index < min_test_days:
    # å¦‚æœæµ‹è¯•é›†å¤ªå°ï¼Œå‘å‰è°ƒæ•´
    test_index = trade_dates - min_test_days
    if test_index <= valid_index:
        # å¦‚æœè¿˜æ˜¯ä¸å¤Ÿï¼Œé‡æ–°åˆ†é…
        available_days = trade_dates - min_train_days
        if available_days >= min_valid_days + min_test_days:
            valid_index = min_train_days
            test_index = trade_dates - min_test_days
        else:
            print(f"é”™è¯¯ï¼šæ•°æ®é›†å¤ªå°({trade_dates}å¤©)ï¼Œæ— æ³•è¿›è¡Œæœ‰æ•ˆçš„è®­ç»ƒ/éªŒè¯/æµ‹è¯•åˆ†å‰²")
            print(f"æœ€å°‘éœ€è¦: {min_train_days + min_valid_days + min_test_days}å¤©")
            exit(1)

print(f"æ ¹æ® {train_ratio:.0%}/{valid_ratio:.0%}/_ åˆ†å‰²æ¯”ä¾‹ (å·²è°ƒæ•´ä¸ºæ»¡è¶³æœ€å°æ•°æ®é›†è¦æ±‚):")
print(f"è®­ç»ƒé›†å¤©æ•° (0-indexed, up to valid_index-1): {valid_index} (å®é™…æ¯”ä¾‹: {valid_index/trade_dates:.1%})")
print(f"éªŒè¯é›†å¤©æ•° (valid_index to test_index-1): {test_index - valid_index} (å®é™…æ¯”ä¾‹: {(test_index-valid_index)/trade_dates:.1%})")
print(f"æµ‹è¯•é›†å¤©æ•° (test_index to end): {trade_dates - test_index} (å®é™…æ¯”ä¾‹: {(trade_dates-test_index)/trade_dates:.1%})")
print(f"éªŒè¯é›†èµ·å§‹ç´¢å¼• (valid_index): {valid_index}")
print(f"æµ‹è¯•é›†èµ·å§‹ç´¢å¼• (test_index): {test_index}")

# æœ€ç»ˆå®‰å…¨æ£€æŸ¥
if valid_index >= test_index or test_index >= trade_dates:
    print(f"é”™è¯¯ï¼šåˆ†å‰²ç´¢å¼•æ— æ•ˆ - valid_index:{valid_index}, test_index:{test_index}, trade_dates:{trade_dates}")
    exit(1)
    
if test_index - valid_index <= 0:
    print(f"é”™è¯¯ï¼šéªŒè¯é›†å¤§å°ä¸º0 - valid_index:{valid_index}, test_index:{test_index}")
    exit(1)
    
if trade_dates - test_index <= 0:
    print(f"é”™è¯¯ï¼šæµ‹è¯•é›†å¤§å°ä¸º0 - test_index:{test_index}, trade_dates:{trade_dates}")
    exit(1)

print("âœ… æ•°æ®é›†åˆ†å‰²æ£€æŸ¥é€šè¿‡")

# --- æ•°æ®æ ‡å‡†åŒ– ---
# åœ¨æ•°æ®é›†åˆ†å‰²åï¼Œä½¿ç”¨è®­ç»ƒé›†æ•°æ®è¿›è¡Œæ ‡å‡†åŒ–
train_eod_data = eod_data[:, :valid_index, :]
data_mean = np.mean(train_eod_data, axis=(0, 1), keepdims=True)
data_std = np.std(train_eod_data, axis=(0, 1), keepdims=True)
data_std[data_std == 0] = 1.0 # é˜²æ­¢é™¤ä»¥é›¶
eod_data = (eod_data - data_mean) / data_std
print("âœ… EODæ•°æ®å·²ä½¿ç”¨è®­ç»ƒé›†çš„å‡å€¼å’Œæ ‡å‡†å·®è¿›è¡Œæ ‡å‡†åŒ–")
# --- ç»“æŸæ•°æ®æ ‡å‡†åŒ– ---

# --- ç»“æŸé€šç”¨é€»è¾‘ ---


# Initialize StockMixer model
calculated_concat_time_dim = 0
original_time_steps = lookback_length # Assuming lookback_length is initial_time_steps for calc
calculated_concat_time_dim += original_time_steps
num_conv_scales_to_add = max(0, scale_factor - 1)
for i in range(num_conv_scales_to_add):
    current_stride = 2**(i + 1)
    ts_after_conv = original_time_steps // current_stride
    if ts_after_conv >= 1:
        calculated_concat_time_dim += ts_after_conv
    else:
        break

# å®šä¹‰æ³¨æ„åŠ›æ¨¡å—ä¸­FFNçš„ç»´åº¦ä¹˜å­
 # å¾®é‡å¢åŠ FFNç»´åº¦ï¼Œä¸“æ³¨æå‡IC
attention_ffn_dim_to_pass = calculated_concat_time_dim * attention_ffn_dim_multiplier

model = StockPredict(
    stocks=stock_num,
    time_steps=lookback_length,
    channels=fea_num,
    scale=scale_factor,
    activation_fn_class=activation_fn_to_pass,
    attention_num_heads=attention_heads,
    attention_hidden_ff_dim=attention_ffn_dim_to_pass,
    attention_dropout_rate=attention_dropout
).to(device)

# --- è®¾ç½®è¡Œä¸šç¨€ç–æ³¨æ„åŠ› ---
print("\n" + "="*60)
print("ğŸ¯ è®¾ç½®åŸºäºè¡Œä¸šçš„ç¨€ç–æ³¨æ„åŠ›æœºåˆ¶")
print("="*60)

# åŠ è½½è‚¡ç¥¨ä»£ç åˆ—è¡¨
ticker_list = load_ticker_list(market_name)

if ticker_list:
    try:
        # è®¾ç½®æ¨¡å‹çš„ç¨€ç–æ³¨æ„åŠ›
        model.setup_sparse_attention(market_name, ticker_list)
        print("âœ… ç¨€ç–æ³¨æ„åŠ›è®¾ç½®æˆåŠŸï¼")
        
        # è·å–ç¨€ç–åº¦ç»Ÿè®¡
        sparse_mask = model.stock_attention_mixer.sparse_attention.industry_mask
        if sparse_mask is not None:
            total_connections = sparse_mask.numel()
            active_connections = sparse_mask.sum().item()
            sparsity = 1 - (active_connections / total_connections)
            theoretical_speedup = total_connections / active_connections
            
            print(f"ğŸ“Š ç¨€ç–æ³¨æ„åŠ›ç»Ÿè®¡:")
            print(f"   è‚¡ç¥¨æ•°é‡: {len(ticker_list)}")
            print(f"   å…¨è¿æ¥æƒ…å†µä¸‹çš„æ³¨æ„åŠ›è®¡ç®—: {total_connections:,} æ¬¡")
            print(f"   ç¨€ç–è¿æ¥ä¸‹çš„æ³¨æ„åŠ›è®¡ç®—: {active_connections:,} æ¬¡")
            print(f"   ç¨€ç–åº¦: {sparsity:.2%}")
            print(f"   ç†è®ºåŠ é€Ÿæ¯”: {theoretical_speedup:.1f}x")
        
    except Exception as e:
        print(f"âŒ ç¨€ç–æ³¨æ„åŠ›è®¾ç½®å¤±è´¥: {e}")
        print("âš ï¸ æ¨¡å‹å°†ä½¿ç”¨æ ‡å‡†çš„å…¨è¿æ¥æ³¨æ„åŠ›æœºåˆ¶")
else:
    print("âš ï¸ æœªæ‰¾åˆ°è‚¡ç¥¨ä»£ç åˆ—è¡¨ï¼Œæ¨¡å‹å°†ä½¿ç”¨æ ‡å‡†çš„å…¨è¿æ¥æ³¨æ„åŠ›æœºåˆ¶")

print("="*60)
print("âœ… æ¨¡å‹åˆå§‹åŒ–å®Œæˆ")
print("="*60 + "\n")
# --- ç»“æŸç¨€ç–æ³¨æ„åŠ›è®¾ç½® ---

optimizer = torch.optim.AdamW(model.parameters(), lr=learning_rate, weight_decay=weight_decay, 
                               betas=(0.9, 0.98), eps=1e-8) # ä½¿ç”¨AdamWä¼˜åŒ–å™¨ï¼Œæ›´å¥½çš„æ­£åˆ™åŒ–å’Œæ”¶æ•›
# ä½¿ç”¨Cosine Annealingå­¦ä¹ ç‡è°ƒåº¦ï¼Œæ›´å¥½çš„æ”¶æ•›æ€§èƒ½
scheduler = torch.optim.lr_scheduler.CosineAnnealingWarmRestarts(
    optimizer, T_0=40, T_mult=2, eta_min=1e-7
)
# --- æ¢¯åº¦ç´¯ç§¯é…ç½® ---
gradient_accumulation_steps = 4  # æ¢¯åº¦ç´¯ç§¯æ­¥æ•°ï¼Œæ¨¡æ‹Ÿæ›´å¤§çš„batch size
best_valid_ic = -np.inf # è®°å½•æœ€ä½³éªŒè¯IC
best_valid_perf = None # è®°å½•æœ€ä½³éªŒè¯é›†æ€§èƒ½
best_test_perf = None # è®°å½•æœ€ä½³éªŒè¯ICå¯¹åº”çš„æµ‹è¯•é›†æ€§èƒ½
best_epoch_num = 0 # è®°å½•æœ€ä½³éªŒè¯æ€§èƒ½å¯¹åº”çš„è½®æ¬¡æ•°
batch_offsets = np.arange(start=0, stop=valid_index, dtype=int) # ç”¨äºè®­ç»ƒæ—¶æ‰“ä¹±æ•°æ®æ‰¹æ¬¡çš„åç§»é‡


def validate(start_index, end_index):
    # åœ¨éªŒè¯é›†æˆ–æµ‹è¯•é›†ä¸Šè¯„ä¼°æ¨¡å‹æ€§èƒ½
    with torch.no_grad(): # å…³é—­æ¢¯åº¦è®¡ç®—
        cur_valid_pred = np.zeros([stock_num, end_index - start_index], dtype=float) # é¢„æµ‹å€¼åˆå§‹åŒ–
        cur_valid_gt = np.zeros([stock_num, end_index - start_index], dtype=float) # çœŸå®å€¼åˆå§‹åŒ–
        cur_valid_mask = np.zeros([stock_num, end_index - start_index], dtype=float) # æ©ç åˆå§‹åŒ–
        loss = 0.
        reg_loss = 0.
        rank_loss = 0.
        # éå†æŒ‡å®šæ—¶é—´æ®µçš„æ•°æ®
        for cur_offset in range(start_index - lookback_length - steps + 1, end_index - lookback_length - steps + 1):
            # è·å–ä¸€ä¸ªæ‰¹æ¬¡çš„æ•°æ® (ä¸å†åŒ…å«volatility_batch)
            data_batch, mask_batch, price_batch, gt_batch = map(
                lambda x: torch.Tensor(x).to(device),
                get_batch(cur_offset)
            )
            prediction = model(data_batch) # æ¨¡å‹å‰å‘ä¼ æ’­
            # è®¡ç®—æŸå¤± (ä¸å†ä¼ é€’volatility_batch)
            cur_loss, cur_reg_loss, cur_rank_loss, cur_rr = get_loss(prediction, gt_batch, price_batch, mask_batch, stock_num, alpha)
            loss += cur_loss.item()
            reg_loss += cur_reg_loss.item()
            rank_loss += cur_rank_loss.item()
            # å­˜å‚¨é¢„æµ‹ç»“æœã€çœŸå®å€¼å’Œæ©ç 
            cur_valid_pred[:, cur_offset - (start_index - lookback_length - steps + 1)] = cur_rr[:, 0].cpu()
            cur_valid_gt[:, cur_offset - (start_index - lookback_length - steps + 1)] = gt_batch[:, 0].cpu()
            cur_valid_mask[:, cur_offset - (start_index - lookback_length - steps + 1)] = mask_batch[:, 0].cpu()
        loss = loss / (end_index - start_index) # è®¡ç®—å¹³å‡æŸå¤±
        reg_loss = reg_loss / (end_index - start_index)
        rank_loss = rank_loss / (end_index - start_index)
        cur_valid_perf = evaluate(cur_valid_pred, cur_valid_gt, cur_valid_mask) # è¯„ä¼°æ€§èƒ½æŒ‡æ ‡
    return loss, reg_loss, rank_loss, cur_valid_perf


def get_batch(offset=None):
    # è·å–ä¸€ä¸ªæ‰¹æ¬¡çš„æ•°æ®ç”¨äºè®­ç»ƒæˆ–éªŒè¯
    if offset is None:
        offset = random.randrange(0, valid_index) # å¦‚æœæœªæŒ‡å®šåç§»é‡ï¼Œåˆ™éšæœºé€‰æ‹©
    seq_len = lookback_length
    # æˆªå–EODæ•°æ®ã€æ©ç æ•°æ®ã€ä»·æ ¼æ•°æ®å’ŒçœŸå®æ”¶ç›Šç‡æ•°æ®
    eod_data_batch = eod_data[:, offset:offset + seq_len, :]
    mask_batch = mask_data[:, offset: offset + seq_len + steps]
    mask_batch = np.min(mask_batch, axis=1) # ç¡®ä¿æ•´ä¸ªåºåˆ—çª—å£å†…æ•°æ®æœ‰æ•ˆ   

    return (
        eod_data_batch,
        np.expand_dims(mask_batch, axis=1), # æ‰©å±•æ©ç ç»´åº¦
        np.expand_dims(price_data[:, offset + seq_len - 1], axis=1), # è·å–åŸºå‡†ä»·æ ¼
        np.expand_dims(gt_data[:, offset + seq_len + steps - 1], axis=1)
        # volatility_batch # ä¸å†è¿”å›æ³¢åŠ¨ç‡æ‰¹æ¬¡
    )


# å¼€å§‹è®­ç»ƒå¾ªç¯
for epoch in range(epochs):
    print("epoch{}##########################################################".format(epoch + 1))
    np.random.shuffle(batch_offsets) # æ¯ä¸ªepochå¼€å§‹æ—¶æ‰“ä¹±æ‰¹æ¬¡é¡ºåº
    tra_loss = 0.0
    tra_reg_loss = 0.0
    tra_rank_loss = 0.0
    
    # æ¢¯åº¦ç´¯ç§¯è®¡æ•°å™¨
    accumulation_steps = 0
    
    # éå†è®­ç»ƒæ•°æ®
    for j in range(valid_index - lookback_length - steps + 1):
        # è·å–ä¸€ä¸ªæ‰¹æ¬¡çš„è®­ç»ƒæ•°æ® (ä¸å†åŒ…å«volatility_batch)
        data_batch, mask_batch, price_batch, gt_batch = map(
            lambda x: torch.Tensor(x).to(device),
            get_batch(batch_offsets[j])
        )
        
        prediction = model(data_batch) # æ¨¡å‹å‰å‘ä¼ æ’­
        # è®¡ç®—æŸå¤± (ä¸å†ä¼ é€’volatility_batch)
        cur_loss, cur_reg_loss, cur_rank_loss, _ = get_loss(prediction, gt_batch, price_batch, mask_batch, stock_num, alpha)
        # æŸå¤±ç¼©æ”¾ä»¥é€‚åº”æ¢¯åº¦ç´¯ç§¯
        cur_loss = cur_loss / gradient_accumulation_steps
        cur_loss.backward() # åå‘ä¼ æ’­
        
        accumulation_steps += 1
        
        # æ¯accumulation_stepsä¸ªbatchæˆ–æœ€åä¸€ä¸ªbatchæ—¶æ›´æ–°å‚æ•°
        if accumulation_steps == gradient_accumulation_steps or j == (valid_index - lookback_length - steps):
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0) # æ·»åŠ æ¢¯åº¦è£å‰ªé˜²æ­¢æ¢¯åº¦çˆ†ç‚¸
            optimizer.step() # æ›´æ–°å‚æ•°
            optimizer.zero_grad() # æ¸…ç©ºæ¢¯åº¦
            accumulation_steps = 0

        tra_loss += cur_loss.item() * gradient_accumulation_steps  # æ¢å¤åŸå§‹æŸå¤±å€¼ç”¨äºè®°å½•
        tra_reg_loss += cur_reg_loss.item()
        tra_rank_loss += cur_rank_loss.item()
    # è®¡ç®—å¹³å‡è®­ç»ƒæŸå¤±
    tra_loss = tra_loss / (valid_index - lookback_length - steps + 1)
    tra_reg_loss = tra_reg_loss / (valid_index - lookback_length - steps + 1)
    tra_rank_loss = tra_rank_loss / (valid_index - lookback_length - steps + 1)
    print('Train : loss:{:.2e}  =  {:.2e} + alpha*{:.2e}'.format(tra_loss, tra_reg_loss, tra_rank_loss))

    # åœ¨éªŒè¯é›†ä¸Šè¯„ä¼°
    val_loss, val_reg_loss, val_rank_loss, val_perf = validate(valid_index, test_index)
    print('Valid : loss:{:.2e}  =  {:.2e} + alpha*{:.2e}'.format(val_loss, val_reg_loss, val_rank_loss))

    # åœ¨æµ‹è¯•é›†ä¸Šè¯„ä¼°
    test_loss, test_reg_loss, test_rank_loss, test_perf = validate(test_index, trade_dates)
    print('Test: loss:{:.2e}  =  {:.2e} + alpha*{:.2e}'.format(test_loss, test_reg_loss, test_rank_loss))

    # å­¦ä¹ ç‡è°ƒåº¦å™¨æ­¥è¿› (CosineAnnealingWarmRestartsæ¯ä¸ªepochè‡ªåŠ¨æ›´æ–°)
    scheduler.step()

    # å¦‚æœå½“å‰éªŒè¯ICæ›´é«˜ï¼Œåˆ™ä¿å­˜æ¨¡å‹æ€§èƒ½
    if val_perf['IC'] > best_valid_ic:
        best_valid_ic = val_perf['IC']
        best_valid_perf = val_perf
        best_test_perf = test_perf # ä¿å­˜æ­¤æ—¶å¯¹åº”çš„æµ‹è¯•é›†æ€§èƒ½
        best_epoch_num = epoch + 1 # æ›´æ–°æœ€ä½³è½®æ¬¡æ•°

    # æ‰“å°éªŒè¯é›†å’Œæµ‹è¯•é›†æ€§èƒ½æŒ‡æ ‡
    print('Valid performance:\n', 'mse:{:.2e}, IC:{:.2e}, RIC:{:.2e}, prec@10:{:.2e}, SR:{:.2e}'.format(val_perf['mse'], val_perf['IC'],
                                                     val_perf['RIC'], val_perf['prec_10'], val_perf['sharpe5']))
    print('Test performance:\n', 'mse:{:.2e}, IC:{:.2e}, RIC:{:.2e}, prec@10:{:.2e}, SR:{:.2e}'.format(test_perf['mse'], test_perf['IC'],
                                                                            test_perf['RIC'], test_perf['prec_10'], test_perf['sharpe5']), '\n\n')

# è®­ç»ƒç»“æŸåæ‰“å°æœ€ä½³éªŒè¯æ€§èƒ½åŠå…¶å¯¹åº”çš„æµ‹è¯•æ€§èƒ½
print("Training finished.")
print(f"Best Validation Performance at Epoch {best_epoch_num} (based on highest validation IC):") # åŸºäºæœ€é«˜éªŒè¯ICçš„æœ€ä½³éªŒè¯æ€§èƒ½
print('mse:{:.2e}, IC:{:.2e}, RIC:{:.2e}, prec@10:{:.2e}, SR:{:.2e}'.format(best_valid_perf['mse'], best_valid_perf['IC'],
                                                                            best_valid_perf['RIC'], best_valid_perf['prec_10'], best_valid_perf['sharpe5']))
print("Corresponding Test Performance:") # å¯¹åº”çš„æµ‹è¯•æ€§èƒ½
print('mse:{:.2e}, IC:{:.2e}, RIC:{:.2e}, prec@10:{:.2e}, SR:{:.2e}'.format(best_test_perf['mse'], best_test_perf['IC'],
                                                                            best_test_perf['RIC'], best_test_perf['prec_10'], best_test_perf['sharpe5']))

# --- å¼€å§‹å†™å…¥Excelçš„é€»è¾‘ ---
if best_valid_perf and best_test_perf and best_epoch_num > 0:
    model_name_for_excel = "StockPredict_SparseAttention"  # ä¿æŒåŸæ¥çš„æ¨¡å‹åç§°
    excel_file_path = '../training_results.xlsx' # ç›¸å¯¹äºsrcç›®å½•
    
    # è·å–ç¨€ç–æ³¨æ„åŠ›ç»Ÿè®¡ä¿¡æ¯
    sparse_attention_info = {
        'enabled': False,
        'sparsity': 0.0,
        'theoretical_speedup': 1.0,
        'industry_groups': 0,
        'stocks_with_industry': 0
    }
    
    try:
        # å°è¯•è·å–ç¨€ç–æ³¨æ„åŠ›ç»Ÿè®¡
        sparse_mask = model.stock_attention_mixer.sparse_attention.industry_mask
        if sparse_mask is not None:
            total_connections = sparse_mask.numel()
            active_connections = sparse_mask.sum().item()
            sparse_attention_info['enabled'] = True
            sparse_attention_info['sparsity'] = 1 - (active_connections / total_connections)
            sparse_attention_info['theoretical_speedup'] = total_connections / active_connections
            
            # ç»Ÿè®¡è¡Œä¸šä¿¡æ¯
            industry_mapping = model.stock_attention_mixer.sparse_attention.industry_mapping
            if industry_mapping:
                sparse_attention_info['stocks_with_industry'] = len(industry_mapping)
                # è®¡ç®—è¡Œä¸šç»„æ•°
                industries = set(industry_mapping.values())
                sparse_attention_info['industry_groups'] = len(industries)
    except Exception as e:
        print(f"âš ï¸ æ— æ³•è·å–ç¨€ç–æ³¨æ„åŠ›ç»Ÿè®¡: {e}")
    
    header = ["Model", "Dataset","Best Epoch",
              "Valid MSE", "Valid IC", "Valid RIC", "Valid Prec@10", "Valid SR",
              "Test MSE", "Test IC", "Test RIC", "Test Prec@10", "Test SR",
              "Lookback Length", "Learning Rate", "Alpha", "Scale Factor", "Activation", 
              "Attention Heads", "Attention Dropout", "Attention FFN Multiplier", "Weight Decay", "Total Epochs",
              "Sparse Attention", "Sparsity (%)", "Theoretical Speedup", "Industry Groups", "Stocks with Industry"
              ]
    
    data_row = [model_name_for_excel, market_name,                
                best_epoch_num,
                best_valid_perf.get('mse', float('nan')), 
                best_valid_perf.get('IC', float('nan')),
                best_valid_perf.get('RIC', float('nan')),
                best_valid_perf.get('prec_10', float('nan')),
                best_valid_perf.get('sharpe5', float('nan')),
                best_test_perf.get('mse', float('nan')),
                best_test_perf.get('IC', 'nan'),
                best_test_perf.get('RIC', float('nan')),
                best_test_perf.get('prec_10', float('nan')),
                best_test_perf.get('sharpe5', float('nan')),
                lookback_length, learning_rate, alpha, scale_factor, activation_str,
                attention_heads, attention_dropout, attention_ffn_dim_multiplier, weight_decay, epochs,
                "Yes" if sparse_attention_info['enabled'] else "No",
                f"{sparse_attention_info['sparsity']*100:.1f}" if sparse_attention_info['enabled'] else "N/A",
                f"{sparse_attention_info['theoretical_speedup']:.1f}" if sparse_attention_info['enabled'] else "N/A",
                sparse_attention_info['industry_groups'] if sparse_attention_info['enabled'] else "N/A",
                sparse_attention_info['stocks_with_industry'] if sparse_attention_info['enabled'] else "N/A"
               ]

    try:
        if not os.path.exists(excel_file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Training Results"
            sheet.append(header)
            # è‡ªåŠ¨è°ƒæ•´åˆ—å®½
            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[get_column_letter(col_idx + 1)].width = length + 2
        else:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active
        
        sheet.append(data_row)
        workbook.save(excel_file_path)
        
        # æ‰“å°ç¨€ç–æ³¨æ„åŠ›ä¿¡æ¯åˆ°ç»“æœ
        print(f"\nğŸ¯ ç¨€ç–æ³¨æ„åŠ›è®­ç»ƒæ€»ç»“:")
        print(f"   æ¨¡å‹åç§°: {model_name_for_excel}")
        print(f"   ç¨€ç–æ³¨æ„åŠ›: {'å¯ç”¨' if sparse_attention_info['enabled'] else 'æœªå¯ç”¨'}")
        if sparse_attention_info['enabled']:
            print(f"   ç¨€ç–åº¦: {sparse_attention_info['sparsity']*100:.1f}%")
            print(f"   ç†è®ºåŠ é€Ÿæ¯”: {sparse_attention_info['theoretical_speedup']:.1f}x")
            print(f"   è¡Œä¸šç»„æ•°: {sparse_attention_info['industry_groups']}")
            print(f"   æœ‰è¡Œä¸šä¿¡æ¯çš„è‚¡ç¥¨: {sparse_attention_info['stocks_with_industry']}")
        
        print(f"Results for {model_name_for_excel} appended to {excel_file_path} (active sheet: {sheet.title})")
    except Exception as e:
        print(f"Error writing {model_name_for_excel} results to Excel: {e}")
else:
    print(f"No best performance data to write to Excel for {model_name_for_excel}.")
# --- ç»“æŸå†™å…¥Excelçš„é€»è¾‘ ---
